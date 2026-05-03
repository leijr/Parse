import openpyxl
import os
import sys
import re
from ReadDBCNew_readlines import readDBC
def DBCtoEXCEL(DBCfilepath,filename,ECU_Name):

    #DBCfilepath=path
    db=readDBC(DBCfilepath)
    #excelfilename=excelName
    #excelfilepath = os.path.abspath(os.curdir)
    #filename=excelfilepath+"\\"+excelfilename+'.xlsx'
    excelfile = openpyxl.Workbook()
    CAN_Tot = excelfile.create_sheet(index=0,title='CAN_Tot')
    CAN_IN = excelfile.create_sheet(index=1,title='CAN_IN')
    CAN_OUT = excelfile.create_sheet(index=2,title='CAN_OUT')
    CAN_Tot['A1']='Sig_Name'
    CAN_Tot['B1']='Min_Value'
    CAN_Tot['C1']='Max_Value'
    CAN_Tot['D1']='Factor'
    CAN_Tot['E1']='offset'
    CAN_Tot['F1']='Main_Frame'
    CAN_Tot['G1']='TX_ECU'
    CAN_Tot['H1']='RX_ECU'
    CAN_Tot['I1']='TX_Cycle_Time'
    CAN_IN['A1']='Sig_Name'
    CAN_IN['B1']='Min_Value'
    CAN_IN['C1']='Max_Value'
    CAN_IN['D1']='Factor'
    CAN_IN['E1']='offset'
    CAN_IN['F1']='Main_Frame'
    CAN_IN['G1']='TX_ECU'
    CAN_IN['H1']='TX_Cycle_Time'
    CAN_OUT['A1']='Sig_Name'
    CAN_OUT['B1']='Min_Value'
    CAN_OUT['C1']='Max_Value'
    CAN_OUT['D1']='Factor'
    CAN_OUT['E1']='offset'
    CAN_OUT['F1']='Main_Frame'
    CAN_OUT['G1']='Main_Frame'
    CAN_OUT['H1']='TX_Cycle_Time'
    Num =1
    Num1 =1
    Num2 =1
    for msgkey in db:
        for sigkey in db[msgkey]['signal']:
            Num +=1
            CAN_Tot['A'+str(Num)] = sigkey
            CAN_Tot['B'+str(Num)] = db[msgkey]['signal'][sigkey]['min']
            CAN_Tot['C'+str(Num)] = db[msgkey]['signal'][sigkey]['max']
            CAN_Tot['D'+str(Num)] = db[msgkey]['signal'][sigkey]['factor']
            CAN_Tot['E'+str(Num)] = db[msgkey]['signal'][sigkey]['offset']
            CAN_Tot['F'+str(Num)] = db[msgkey]['txecu']
            CAN_Tot['G'+str(Num)] = msgkey
            CAN_Tot['H'+str(Num)] = ','.join(db[msgkey]['signal'][sigkey]['rxecu'])
            CAN_Tot['I' + str(Num)] = db[msgkey]['TxCycTime']
        if db[msgkey]['txecu'] == ECU_Name:
        ######################################################################################
            for sigkey in db[msgkey]['signal']:
                Num1 +=1
                CAN_OUT['A'+str(Num1)] = sigkey
                CAN_OUT['B'+str(Num1)] = db[msgkey]['signal'][sigkey]['min']
                CAN_OUT['C'+str(Num1)] = db[msgkey]['signal'][sigkey]['max']
                CAN_OUT['D'+str(Num1)] = db[msgkey]['signal'][sigkey]['factor']
                CAN_OUT['E'+str(Num1)] = db[msgkey]['signal'][sigkey]['offset']
                CAN_OUT['F'+str(Num1)] = msgkey
                CAN_OUT['G' + str(Num1)] = db[msgkey]['txecu']
                CAN_OUT['H' + str(Num1)] = db[msgkey]['TxCycTime']
        elif db[msgkey]['txecu'] != ECU_Name :
            for sigkey in db[msgkey]['signal']:
                if ECU_Name in db[msgkey]['signal'][sigkey]['rxecu']:
                    Num2 +=1
                    CAN_IN['A' + str(Num2)] = sigkey
                    CAN_IN['B' + str(Num2)] = db[msgkey]['signal'][sigkey]['min']
                    CAN_IN['C' + str(Num2)] = db[msgkey]['signal'][sigkey]['max']
                    CAN_IN['D' + str(Num2)] = db[msgkey]['signal'][sigkey]['factor']
                    CAN_IN['E' + str(Num2)] = db[msgkey]['signal'][sigkey]['offset']
                    CAN_IN['F' + str(Num2)] = msgkey
                    CAN_IN['G' + str(Num2)] = db[msgkey]['txecu']
                    CAN_IN['H' + str(Num2)] = db[msgkey]['TxCycTime']

    excelfile.save(filename)
    excelfile.close()
