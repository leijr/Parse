"""CAN signal matching and DBC-to-Excel conversion.

Matches CAN interface signals between DBC definitions and
external Excel interface files.
"""

import os
import openpyxl

from .dbc_reader import read_dbc


def dbc_to_excel(dbc_path: str, output_path: str, ecu_name: str):
    """Convert DBC file to Excel with CAN_Tot, CAN_IN, CAN_OUT sheets.

    Args:
        dbc_path: Path to .dbc file
        output_path: Path for output .xlsx file
        ecu_name: Selected ECU name for TX/RX classification
    """
    db = read_dbc(dbc_path)
    wb = openpyxl.Workbook()

    sheets = {
        "CAN_Tot": {"idx": 0, "headers": [
            "Sig_Name", "Min_Value", "Max_Value", "Factor", "offset",
            "Main_Frame", "TX_ECU", "RX_ECU", "TX_Cycle_Time"
        ]},
        "CAN_IN": {"idx": 1, "headers": [
            "Sig_Name", "Min_Value", "Max_Value", "Factor", "offset",
            "Main_Frame", "TX_ECU", "TX_Cycle_Time"
        ]},
        "CAN_OUT": {"idx": 2, "headers": [
            "Sig_Name", "Min_Value", "Max_Value", "Factor", "offset",
            "Main_Frame", "TX_ECU", "TX_Cycle_Time"
        ]},
    }

    for name, info in sheets.items():
        if info["idx"] == 0:
            ws = wb.active
            ws.title = name
        else:
            ws = wb.create_sheet(title=name)
        for col, hdr in enumerate(info["headers"], 1):
            ws.cell(row=1, column=col, value=hdr)

    # Populate
    row_all = 1
    row_in = 1
    row_out = 1

    for msg_name, msg_data in db.items():
        tx_ecu = msg_data["txecu"]
        is_tx = (tx_ecu == ecu_name)

        for sig_name, sig_data in msg_data["signal"].items():
            row_all += 1
            ws_tot = wb["CAN_Tot"]
            ws_tot.cell(row=row_all, column=1, value=sig_name)
            ws_tot.cell(row=row_all, column=2, value=sig_data["min"])
            ws_tot.cell(row=row_all, column=3, value=sig_data["max"])
            ws_tot.cell(row=row_all, column=4, value=sig_data["factor"])
            ws_tot.cell(row=row_all, column=5, value=sig_data["offset"])
            ws_tot.cell(row=row_all, column=6, value=msg_name)
            ws_tot.cell(row=row_all, column=7, value=tx_ecu)
            ws_tot.cell(row=row_all, column=8,
                        value=",".join(sig_data.get("rxecu", [])))
            ws_tot.cell(row=row_all, column=9,
                        value=msg_data.get("TxCycTime", ""))

            if is_tx:
                row_out += 1
                ws = wb["CAN_OUT"]
                ws.cell(row=row_out, column=1, value=sig_name)
                ws.cell(row=row_out, column=2, value=sig_data["min"])
                ws.cell(row=row_out, column=3, value=sig_data["max"])
                ws.cell(row=row_out, column=4, value=sig_data["factor"])
                ws.cell(row=row_out, column=5, value=sig_data["offset"])
                ws.cell(row=row_out, column=6, value=msg_name)
                ws.cell(row=row_out, column=7, value=tx_ecu)
                ws.cell(row=row_out, column=8,
                        value=msg_data.get("TxCycTime", ""))

            if ecu_name in sig_data.get("rxecu", []):
                row_in += 1
                ws = wb["CAN_IN"]
                ws.cell(row=row_in, column=1, value=sig_name)
                ws.cell(row=row_in, column=2, value=sig_data["min"])
                ws.cell(row=row_in, column=3, value=sig_data["max"])
                ws.cell(row=row_in, column=4, value=sig_data["factor"])
                ws.cell(row=row_in, column=5, value=sig_data["offset"])
                ws.cell(row=row_in, column=6, value=msg_name)
                ws.cell(row=row_in, column=7, value=tx_ecu)
                ws.cell(row=row_in, column=8,
                        value=msg_data.get("TxCycTime", ""))

    wb.save(output_path)
    wb.close()


def _match_and_filter(excel_path: str, dbc_excel_path: str,
                      sheet_name: str, match_rules: list[tuple],
                      output_path: str):
    """Generic CAN signal matching engine.

    Args:
        excel_path: Interface Excel file path
        dbc_excel_path: DBC-converted Excel file path
        sheet_name: Sheet name in interface Excel to match
        match_rules: List of (prefix, suffix, target_col, header) tuples
        output_path: Output Excel file path
    """
    wb_if = openpyxl.load_workbook(excel_path)
    wb_dbc = openpyxl.load_workbook(dbc_excel_path)

    ws_if = wb_if[sheet_name]
    ws_dbc = wb_dbc["CAN_Tot"]

    # Build lookup: signal name -> row number in CAN_Tot
    dbc_signals: dict[str, int] = {}
    for row in range(2, ws_dbc.max_row + 1):
        sig_name = ws_dbc.cell(row=row, column=1).value
        if sig_name:
            dbc_signals[sig_name] = row

    matched_rows: list[int] = []
    for row in range(2, ws_if.max_row + 1):
        cell_val = str(ws_if.cell(row=row, column=1).value or "")
        matched = False
        for prefix, suffix, col, header in match_rules:
            pattern = f"{prefix}{{}}{suffix}"
            for sig_name in dbc_signals:
                candidate = pattern.format(sig_name)
                if cell_val == candidate:
                    ws_if.cell(row=row, column=ord(col) - ord("A") + 1,
                               value=sig_name)
                    matched = True
                    matched_rows.append(row)
                    break
            if matched:
                break

    # Add header if provided
    for _, _, col, header in match_rules:
        if header:
            col_idx = ord(col) - ord("A") + 1
            ws_if.cell(row=1, column=col_idx, value=header)

    # Delete unmatched rows (from bottom up)
    all_rows = set(range(2, ws_if.max_row + 1))
    unmatched = sorted(all_rows - set(matched_rows), reverse=True)
    for row in unmatched:
        ws_if.delete_rows(row)

    wb_if.save(output_path)
    wb_if.close()
    wb_dbc.close()


def match_can_in(excel_path: str, dbc_excel_path: str,
                 sheet_name: str, output_path: str):
    """Match CAN_IN signals.

    Looks for ic_{signal_name} pattern in the interface Excel.
    """
    _match_and_filter(
        excel_path, dbc_excel_path, sheet_name,
        [("ic_", "", "I", "Meas")],
        output_path,
    )


def match_can_out(excel_path: str, dbc_excel_path: str,
                  sheet_name: str, output_path: str):
    """Match CAN_OUT signals.

    Looks for oc_{signal_name}Man_C and oc_{signal_name}EnaMan_C patterns.
    """
    _match_and_filter(
        excel_path, dbc_excel_path, sheet_name,
        [
            ("oc_", "Man_C", "I", "Calibration"),
            ("oc_", "EnaMan_C", "J", "Debug"),
        ],
        output_path,
    )
