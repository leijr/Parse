import openpyxl


def _match_and_filter(excel_file, dbc_excel, sheet_name, match_rules, save_filename):
    """Generic CAN signal matching and row filtering.

    Args:
        excel_file: Path to the interface Excel file
        dbc_excel: Path to the DBC-generated Excel file
        sheet_name: Sheet name to read from dbc_excel ('CAN_IN' or 'CAN_OUT')
        match_rules: List of (prefix, suffix, col_letter, col_header) tuples
        save_filename: Output Excel path
    """
    interface_wb = openpyxl.load_workbook(excel_file)
    dbc_wb = openpyxl.load_workbook(dbc_excel)

    dbc_sheet = dbc_wb[sheet_name]
    interface_sheet = interface_wb[sheet_name]

    for col_letter, col_header in [(r[2], r[3]) for r in match_rules]:
        dbc_sheet[f'{col_letter}1'] = col_header

    dbc_signals = [data[0].value for data in list(dbc_sheet.rows)[1:]]
    interface_signals = [data[0].value for data in list(interface_sheet.rows)[1:]]

    for i, dbc_sig in enumerate(dbc_signals):
        for prefix, suffix, col_letter, _ in match_rules:
            match_str = f'{prefix}{dbc_sig}{suffix}'
            for interface_sig in interface_signals:
                if match_str == interface_sig:
                    dbc_sheet[f'{col_letter}{i + 2}'] = interface_sig
                    break

    # Delete rows that have no matches
    for i in range(2, dbc_sheet.max_row + 1):
        j = 1
        n = 0
        while j in range(1, dbc_sheet.max_column + 1):
            if dbc_sheet.cell(i, j).value is None:
                dbc_sheet.delete_rows(i)
                j = 0
                n += 1
                if i + n >= dbc_sheet.max_row:
                    break
            j += 1

    dbc_wb.save(save_filename)
    dbc_wb.close()


def DBC_CANIN(CANIN_Excel, DBC_Excel, sheetname, CANIN_SAVE_filename):
    _match_and_filter(
        CANIN_Excel, DBC_Excel, sheetname,
        match_rules=[('ic_', '', 'I', 'Meas')],
        save_filename=CANIN_SAVE_filename,
    )


def DBC_CANOUT(CANOUT_Excel, DBC_Excel, sheetname, CANOUT_SAVE_filename):
    _match_and_filter(
        CANOUT_Excel, DBC_Excel, sheetname,
        match_rules=[
            ('oc_', 'Man_C', 'I', 'Calibration'),
            ('oc_', 'EnaMan_C', 'J', 'Debug'),
        ],
        save_filename=CANOUT_SAVE_filename,
    )
